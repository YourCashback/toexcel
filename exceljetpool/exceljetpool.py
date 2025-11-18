import requests
from bs4 import BeautifulSoup
import openpyxl
import os
import win32com.client as win32
import tkinter as tk
from tkinter import ttk, messagebox, scrolledtext
import threading
import sys
from pathlib import Path
import pythoncom

class PriceParserApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Парсер цен товаров")
        self.root.geometry("600x400")
        self.root.resizable(True, True)
        
        # Определяем путь для сохранения Excel файла (рядом с EXE)
        self.excel_file = self.get_excel_file_path()
        
        # Флаг для корректного завершения
        self.is_closing = False
        self.active_threads = []
        
        # Обработка закрытия окна
        self.root.protocol("WM_DELETE_WINDOW", self.safe_close)
        
        # Создаем и размещаем элементы интерфейса
        self.create_widgets()
        
    def get_excel_file_path(self):
        """Определяет путь для сохранения Excel файла рядом с EXE"""
        if getattr(sys, 'frozen', False):
            base_path = Path(sys.executable).parent
        else:
            base_path = Path(__file__).parent
        
        excel_path = base_path / "prices.xlsx"
        return str(excel_path)
    
    def create_widgets(self):
        # Заголовок
        title_label = ttk.Label(self.root, text="Парсер цен товаров", font=("Calibri", 16, "bold"))
        title_label.pack(pady=10)
        
        # Информация о пути сохранения
        path_info = ttk.Label(self.root, text=f"Файл будет сохранен: {os.path.basename(self.excel_file)}", 
                             font=("Calibri", 9), foreground="blue")
        path_info.pack(pady=2)
        
        # Описание
        desc_label = ttk.Label(self.root, text="Введите ссылку на товар с вашего сайта:", font=("Calibri", 10))
        desc_label.pack(pady=5)
        
        # Поле для ввода ссылки с кнопкой вставки
        self.url_frame = ttk.Frame(self.root)
        self.url_frame.pack(pady=10, padx=20, fill=tk.X)
        
        self.url_label = ttk.Label(self.url_frame, text="Ссылка:", font=("Calibri", 10))
        self.url_label.pack(side=tk.LEFT)
        
        self.url_entry = ttk.Entry(self.url_frame, width=50, font=("Calibri", 10))
        self.url_entry.pack(side=tk.LEFT, padx=(10, 5), fill=tk.X, expand=True)
        self.url_entry.bind('<Return>', lambda event: self.add_to_table())
        
        # Кнопка вставки из буфера обмена
        self.paste_button = ttk.Button(self.url_frame, text="Вставить", 
                                      command=self.paste_from_clipboard, width=10)
        self.paste_button.pack(side=tk.LEFT, padx=5)
        
        # Кнопки действий
        self.button_frame = ttk.Frame(self.root)
        self.button_frame.pack(pady=15)
        
        self.add_button = ttk.Button(self.button_frame, text="Добавить в таблицу", 
                                   command=self.add_to_table, width=20)
        self.add_button.pack(side=tk.LEFT, padx=5)
        
        self.clear_button = ttk.Button(self.button_frame, text="Очистить поле", 
                                     command=self.clear_field, width=15)
        self.clear_button.pack(side=tk.LEFT, padx=5)
        
        self.exit_button = ttk.Button(self.button_frame, text="Выйти", 
                                    command=self.safe_close, width=15)
        self.exit_button.pack(side=tk.LEFT, padx=5)
        
        # Область для вывода логов
        self.log_frame = ttk.LabelFrame(self.root, text="Лог выполнения", padding=10)
        self.log_frame.pack(pady=10, padx=20, fill=tk.BOTH, expand=True)
        
        self.log_text = scrolledtext.ScrolledText(self.log_frame, height=15, width=70, 
                                                font=("Consolas", 9))
        self.log_text.pack(fill=tk.BOTH, expand=True)
        self.log_text.config(state=tk.DISABLED)
        
        # Статус бар
        self.status_var = tk.StringVar()
        self.status_var.set("Готов к работе")
        self.status_bar = ttk.Label(self.root, textvariable=self.status_var, relief=tk.SUNKEN, anchor=tk.W)
        self.status_bar.pack(side=tk.BOTTOM, fill=tk.X)
        
        # Контекстное меню для поля ввода
        self.create_context_menu()
        
        # Фокус на поле ввода
        self.url_entry.focus()
        
        # Создаем Excel файл при запуске, если его нет
        self.create_excel_file_if_not_exists()
    
    def create_excel_file_if_not_exists(self):
        """Создает Excel файл при запуске, если он не существует"""
        if not os.path.exists(self.excel_file):
            try:
                wb = openpyxl.Workbook()
                ws = wb.active
                # Создаем заголовки
                ws['A1'] = "Наименование товара"
                ws['B1'] = "Цена"
                # Устанавливаем шрифт для заголовков
                ws['A1'].font = openpyxl.styles.Font(name='Calibri', size=14, bold=True)
                ws['B1'].font = openpyxl.styles.Font(name='Calibri', size=14, bold=True)
                wb.save(self.excel_file)
                self.log_message(f"Создан новый файл: {os.path.basename(self.excel_file)}")
            except Exception as e:
                self.log_message(f"Ошибка при создании файла: {e}")
    
    def create_context_menu(self):
        """Создает контекстное меню для поля ввода"""
        self.context_menu = tk.Menu(self.root, tearoff=0)
        self.context_menu.add_command(label="Вставить", command=self.paste_from_clipboard)
        self.context_menu.add_command(label="Вырезать", command=self.cut_text)
        self.context_menu.add_command(label="Копировать", command=self.copy_text)
        self.context_menu.add_separator()
        self.context_menu.add_command(label="Очистить", command=self.clear_field)
        
        # Привязываем контекстное меню к полю ввода
        self.url_entry.bind("<Button-3>", self.show_context_menu)
    
    def show_context_menu(self, event):
        """Показывает контекстное меню"""
        if self.is_closing:
            return
        try:
            self.context_menu.tk_popup(event.x_root, event.y_root)
        finally:
            self.context_menu.grab_release()
    
    def safe_close(self):
        """Безопасное закрытие приложения"""
        if self.is_closing:
            return
            
        self.is_closing = True
        
        # Проверяем активные потоки
        if self.active_threads:
            active_count = sum(1 for thread in self.active_threads if thread.is_alive())
            if active_count > 0:
                result = messagebox.askyesno(
                    "Подождите", 
                    f"В настоящее время выполняется {active_count} операций. Вы уверены, что хотите выйти?"
                )
                if not result:
                    self.is_closing = False
                    return
        
        # Блокируем все кнопки
        self.add_button.config(state=tk.DISABLED)
        self.clear_button.config(state=tk.DISABLED)
        self.paste_button.config(state=tk.DISABLED)
        self.exit_button.config(state=tk.DISABLED)
        
        self.log_message("Завершение работы приложения...")
        self.update_status("Завершение...")
        
        # Даем время на завершение операций
        self.root.after(100, self.force_close)
    
    def force_close(self):
        """Принудительное закрытие приложения"""
        try:
            # Закрываем все соединения с Excel
            try:
                win32.Dispatch("Excel.Application").Quit()
            except:
                pass
                
            # Закрываем окно
            self.root.quit()
            self.root.destroy()
            
        except Exception as e:
            print(f"Ошибка при закрытии: {e}")
        
        # Гарантированное завершение процесса
        import time
        time.sleep(0.5)
        os._exit(0)
    
    def paste_from_clipboard(self):
        """Вставляет текст из буфера обмена в поле ввода"""
        if self.is_closing:
            return
        try:
            self.url_entry.delete(0, tk.END)
            clipboard_text = self.root.clipboard_get()
            self.url_entry.insert(0, clipboard_text)
            self.log_message("Текст из буфера обмена вставлен")
        except Exception as e:
            self.log_message(f"Ошибка при вставке из буфера обмена: {e}")
            messagebox.showerror("Ошибка", "Не удалось вставить текст из буфера обмена")
    
    def cut_text(self):
        """Вырезает выделенный текст"""
        if self.is_closing:
            return
        try:
            self.url_entry.event_generate("<<Cut>>")
        except:
            pass
    
    def copy_text(self):
        """Копирует выделенный текст"""
        if self.is_closing:
            return
        try:
            self.url_entry.event_generate("<<Copy>>")
        except:
            pass
    
    def log_message(self, message):
        """Добавляет сообщение в лог"""
        if self.is_closing:
            return
        try:
            self.log_text.config(state=tk.NORMAL)
            self.log_text.insert(tk.END, message + "\n")
            self.log_text.see(tk.END)
            self.log_text.config(state=tk.DISABLED)
            self.root.update_idletasks()
        except:
            pass
    
    def update_status(self, message):
        """Обновляет статус бар"""
        if self.is_closing:
            return
        try:
            self.status_var.set(message)
            self.root.update_idletasks()
        except:
            pass
    
    def clear_field(self):
        """Очищает поле ввода"""
        if self.is_closing:
            return
        self.url_entry.delete(0, tk.END)
        self.url_entry.focus()
        self.log_message("Поле ввода очищено")
    
    def parse_my_site(self, url):
        """Парсит название и цену товара с вашего сайта"""
        if self.is_closing:
            return None, None
            
        try:
            self.update_status(f"Парсим сайт: {url}")
            self.log_message(f"Начинаем парсинг: {url}")
            
            headers = {
                'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'
            }
            response = requests.get(url, headers=headers, timeout=10)
            response.raise_for_status()
            
            soup = BeautifulSoup(response.content, 'html.parser')
            
            # Поиск названия товара в атрибуте data-product-name
            product_element = soup.find(attrs={"data-product-name": True})
            if product_element:
                product_name = product_element.get('data-product-name', '').strip()
                self.log_message(f"Найдено название: {product_name}")
            else:
                product_name = "Не удалось найти название"
                self.log_message("Не удалось найти название товара")
            
            # Поиск цены товара
            price_tag = soup.find('meta', {'property': 'product:price:amount'})
            if price_tag:
                price = price_tag.get('content', '').strip()
                self.log_message(f"Найдена цена: {price} руб.")
            else:
                price = "0"
                self.log_message("Не удалось найти цену")
            
            return product_name, price
            
        except Exception as e:
            error_msg = f"Ошибка при парсинге вашего сайта: {e}"
            self.log_message(error_msg)
            return None, None
    
    def find_first_empty_row_in_column_a(self, ws, method='win32com'):
        """Находит первую действительно пустую строку в столбце A"""
        if method == 'win32com':
            # Для win32com - проверяем только столбец A
            row = 2  # Начинаем со второй строки (после заголовков)
            # Ищем до максимально возможного количества строк в Excel (1,048,576)
            while row <= 1048576:
                try:
                    cell_value = ws.Cells(row, 1).Value
                    if cell_value is None or str(cell_value).strip() == "":
                        return row
                    row += 1
                except:
                    return row
            return row
        else:
            # Для openpyxl - проверяем только столбец A
            # Используем max_row для определения заполненных строк, но продолжаем поиск дальше
            max_checked = max(ws.max_row + 1, 2)  # Начинаем с max_row+1 или со 2-й строки
            for row in range(2, max_checked + 1000):  # Проверяем на 1000 строк дальше max_row
                try:
                    cell_value = ws.cell(row=row, column=1).value
                    if cell_value is None or str(cell_value).strip() == "":
                        return row
                except:
                    return row
            return max_checked + 1000
    
    def update_excel_with_win32com(self, product_name, my_price):
        """Обновляет Excel файл с ценами через win32com"""
        if self.is_closing:
            return False
            
        # Инициализируем COM для этого потока
        pythoncom.CoInitialize()
        
        excel = None
        wb = None
        
        try:
            self.update_status("Обновляем Excel файл...")
            
            # Используем self.excel_file
            abs_path = os.path.abspath(self.excel_file)
            
            # Подключаемся к Excel
            excel = win32.Dispatch("Excel.Application")
            excel.Visible = False  # Скрываем Excel
            
            # Пытаемся найти открытую книгу
            try:
                wb = excel.Workbooks.Open(abs_path)
            except Exception as e:
                self.log_message(f"Не удалось открыть файл через win32com: {e}")
                return False
            
            ws = wb.ActiveSheet
            
            # Ищем товар в столбце A (наименования), начиная со 2-й строки
            found = False
            row = 2  # Начинаем со второй строки (после заголовков)
            
            # Ищем только до первой пустой строки в столбце A
            while row <= 1048576:  # Максимальное количество строк в Excel
                try:
                    cell_value = ws.Cells(row, 1).Value
                    if cell_value is None or str(cell_value).strip() == "":
                        # Дошли до пустой строки - останавливаем поиск
                        break
                        
                    if self.is_closing:
                        if wb:
                            wb.Close(False)
                        if excel:
                            excel.Quit()
                        return False
                        
                    if ws.Cells(row, 1).Value == product_name:
                        # Обновляем цену в столбце B (наш сайт)
                        ws.Cells(row, 2).Value = my_price
                        ws.Cells(row, 2).Font.Name = "Calibri"
                        ws.Cells(row, 2).Font.Size = 18
                        
                        found = True
                        self.log_message(f"Обновлена цена для: {product_name}")
                        break
                    row += 1
                except:
                    break
            
            # Если товар не найден, добавляем в первую пустую строку столбца A
            if not found and not self.is_closing:
                # Находим первую действительно пустую строку в столбце A
                new_row = self.find_first_empty_row_in_column_a(ws, 'win32com')
                ws.Cells(new_row, 1).Value = product_name
                ws.Cells(new_row, 2).Value = my_price
                ws.Cells(new_row, 1).Font.Name = "Calibri"
                ws.Cells(new_row, 1).Font.Size = 18
                ws.Cells(new_row, 2).Font.Name = "Calibri"
                ws.Cells(new_row, 2).Font.Size = 18
                self.log_message(f"Добавлен новый товар в строку {new_row}: {product_name}")
            
            # Сохраняем файл
            if not self.is_closing:
                wb.Save()
                self.log_message(f"Файл {os.path.basename(self.excel_file)} успешно обновлен!")
                return True
            else:
                if wb:
                    wb.Close(False)
                return False
            
        except Exception as e:
            error_msg = f"Ошибка при работе с Excel через win32com: {e}"
            self.log_message(error_msg)
            return False
        finally:
            # Всегда закрываем Excel и освобождаем COM
            try:
                if wb:
                    wb.Close()
            except:
                pass
            try:
                if excel:
                    excel.Quit()
            except:
                pass
            # Освобождаем COM
            pythoncom.CoUninitialize()
    
    def update_excel_with_openpyxl(self, product_name, my_price):
        """Обновляет Excel файл с ценами через openpyxl"""
        if self.is_closing:
            return False
            
        try:
            self.update_status("Обновляем Excel файл (альтернативный метод)...")
            
            wb = openpyxl.load_workbook(self.excel_file)
            ws = wb.active
            
            # Ищем товар в столбце A (наименования), начиная со 2-й строки
            found = False
            # Ищем только в строках, где есть данные в столбце A
            max_row_to_check = ws.max_row + 1000  # Проверяем на 1000 строк дальше текущего max_row
            
            for row in range(2, max_row_to_check + 1):
                try:
                    cell_value = ws.cell(row=row, column=1).value
                    if cell_value is None or str(cell_value).strip() == "":
                        # Дошли до пустой строки - останавливаем поиск
                        break
                        
                    if self.is_closing:
                        return False
                        
                    if ws.cell(row=row, column=1).value == product_name:
                        ws.cell(row=row, column=2).value = my_price
                        ws.cell(row=row, column=2).font = openpyxl.styles.Font(name='Calibri', size=18)
                        found = True
                        self.log_message(f"Обновлена цена для: {product_name}")
                        break
                except:
                    break
            
            # Если товар не найден, добавляем в первую пустую строку столбца A
            if not found and not self.is_closing:
                # Находим первую действительно пустую строку в столбце A
                new_row = self.find_first_empty_row_in_column_a(ws, 'openpyxl')
                ws.cell(row=new_row, column=1).value = product_name
                ws.cell(row=new_row, column=2).value = my_price
                ws.cell(row=new_row, column=1).font = openpyxl.styles.Font(name='Calibri', size=18)
                ws.cell(row=new_row, column=2).font = openpyxl.styles.Font(name='Calibri', size=18)
                self.log_message(f"Добавлен новый товар в строку {new_row}: {product_name}")
            
            if not self.is_closing:
                wb.save(self.excel_file)
                self.log_message(f"Файл {os.path.basename(self.excel_file)} успешно обновлен!")
                return True
            else:
                return False
            
        except Exception as e:
            error_msg = f"Ошибка при работе с Excel через openpyxl: {e}"
            self.log_message(error_msg)
            if "Permission denied" in str(e):
                self.log_message("Файл открыт в Excel. Пытаемся использовать альтернативный метод...")
                return self.update_excel_with_win32com(product_name, my_price)
            return False
    
    def update_excel(self, product_name, my_price):
        """Основная функция обновления Excel"""
        if self.is_closing:
            return False
            
        try:
            # Сначала пытаемся использовать win32com
            if self.update_excel_with_win32com(product_name, my_price):
                return True
        except Exception as e:
            self.log_message(f"win32com не доступен: {e}. Пробуем openpyxl...")
        
        # Если win32com не работает, используем openpyxl
        return self.update_excel_with_openpyxl(product_name, my_price)
    
    def add_to_table_thread(self):
        """Функция для выполнения в отдельном потоке"""
        thread_id = threading.current_thread().ident
        self.active_threads.append(threading.current_thread())
        
        try:
            if self.is_closing:
                return
                
            url = self.url_entry.get().strip()
            
            if not url:
                if not self.is_closing:
                    messagebox.showwarning("Внимание", "Введите ссылку на товар!")
                return
            
            if not url.startswith('http'):
                if not self.is_closing:
                    messagebox.showwarning("Внимание", "Неверный формат ссылки!")
                return
            
            # Блокируем кнопки на время выполнения
            if not self.is_closing:
                self.add_button.config(state=tk.DISABLED)
                self.clear_button.config(state=tk.DISABLED)
                self.paste_button.config(state=tk.DISABLED)
            
            try:
                # Парсинг данных с вашего сайта
                product_name, my_price = self.parse_my_site(url)
                
                if product_name and my_price and not self.is_closing:
                    # Подтверждение действия
                    result = messagebox.askyesno("Подтверждение", 
                                               f"Найдено: {product_name}\nЦена: {my_price} руб.\n\nДобавить в таблицу?")
                    
                    if result and not self.is_closing:
                        if self.update_excel(product_name, my_price):
                            if not self.is_closing:
                                messagebox.showinfo("Успех", "Данные успешно добавлены в таблицу!")
                                self.clear_field()
                        else:
                            if not self.is_closing:
                                messagebox.showerror("Ошибка", "Не удалось добавить данные в таблицу!")
                elif not self.is_closing:
                    messagebox.showerror("Ошибка", "Не удалось получить данные о товаре!")
                    
            except Exception as e:
                if not self.is_closing:
                    messagebox.showerror("Ошибка", f"Произошла ошибка: {e}")
            finally:
                # Разблокируем кнопки
                if not self.is_closing:
                    self.add_button.config(state=tk.NORMAL)
                    self.clear_button.config(state=tk.NORMAL)
                    self.paste_button.config(state=tk.NORMAL)
                    self.update_status("Готов к работе")
                
        finally:
            # Удаляем поток из списка активных
            self.active_threads = [t for t in self.active_threads if t.ident != thread_id]
    
    def add_to_table(self):
        """Запускает добавление в таблицу в отдельном потоке"""
        if self.is_closing:
            return
            
        thread = threading.Thread(target=self.add_to_table_thread)
        thread.daemon = True
        thread.start()

def main():
    # Создаем главное окно
    root = tk.Tk()
    
    # Настраиваем стиль
    style = ttk.Style()
    style.theme_use('vista')
    
    # Создаем приложение
    app = PriceParserApp(root)
    
    try:
        # Запускаем главный цикл
        root.mainloop()
    except Exception as e:
        print(f"Ошибка в главном цикле: {e}")
    finally:
        # Гарантированное завершение
        os._exit(0)

if __name__ == "__main__":
    main()